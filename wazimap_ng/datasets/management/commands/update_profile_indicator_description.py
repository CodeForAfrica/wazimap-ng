import openpyxl
from django.core.management.base import BaseCommand
from wazimap_ng.profile.models import Profile, ProfileIndicator, IndicatorSubcategory
from wazimap_ng.datasets.models import Geography

class Command(BaseCommand):
    help = 'Update ProfileIndicator descriptions from Excel file.'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help="Path to the Excel file.")
        parser.add_argument('profile_name', type=str, help="Name of the profile to update.")

    def handle(self, *args, **options):
        file_path = options['file_path']
        profile_name = options['profile_name']

        try:
            profile = Profile.objects.get(name=profile_name)
        except Profile.DoesNotExist:
            self.stdout.write(self.style.ERROR(f"Profile with name '{profile_name}' does not exist."))
            return

        try:
            subcategory = IndicatorSubcategory.objects.get(name="Decadal Temperature Variation", category__profile=profile)
        except IndicatorSubcategory.DoesNotExist:
            self.stdout.write(self.style.ERROR(f"Subcategory 'Decadal Temperature Variation' does not exist for profile '{profile_name}'."))
            return

        # Load the Excel file
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error opening the Excel file: {e}"))
            return

        # Iterate through each worksheet in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.stdout.write(self.style.SUCCESS(f"Processing sheet: {sheet_name}"))

            # Loop through the rows in each worksheet, starting from row 2 (to skip the header)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                region_code = row[0]  # Column A: Region Code (e.g., KE.BA)
                description_text = row[1]  # Column B: Text

                if not region_code or not description_text:
                    self.stdout.write(self.style.WARNING(f"Skipping row with empty region code or description in {sheet_name}."))
                    continue

                try:
                    geography = Geography.objects.get(code=region_code.strip())
                    profile_indicator = ProfileIndicator.objects.get(profile=profile,subcategory=subcategory, indicator__name__icontains=geography.code)
                    profile_indicator.description = description_text
                    profile_indicator.save()

                    self.stdout.write(self.style.SUCCESS(f"Updated description for {profile_indicator.label} in region {region_code}."))
                except Geography.DoesNotExist:
                    self.stdout.write(self.style.WARNING(f"Geography with code '{region_code}' does not exist."))
                except ProfileIndicator.DoesNotExist:
                    self.stdout.write(self.style.WARNING(f"ProfileIndicator for region code '{region_code}' does not exist."))
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error processing region code '{region_code}': {e}"))

        self.stdout.write(self.style.SUCCESS("Finished updating ProfileIndicator descriptions."))

